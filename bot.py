import os
import logging
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, ContextTypes, filters
import openpyxl
import re

# === Настройки ===
BOT_TOKEN = os.getenv("BOT_TOKEN")
EXCEL_FILE = "номера_заглавные.xlsx"  # Путь к Excel файлу с номерами
MOSREG_FILE = "Московская область.txt"
MOTO_FILE = "moto_numbers.txt"
TRAILER_FILE = "trailer_numbers.txt"
MOSCOW_FILE = "270315af-8756-4519-b3cf-88fac83dbc0b.txt"
DEFAULT_PAGE_SIZE = 30

def ru_to_lat(text):
    """Преобразует русские буквы в латинские аналоги"""
    repl = str.maketrans("АВЕКМНОРСТУХ", "ABEKMHOPCTYX")
    return text.translate(repl)

def extract_letters_from_number(number):
    """Извлекает только буквы из номера и преобразует в заглавные"""
    # Извлекаем все буквы (русские и латинские)
    letters = re.findall(r"[А-ЯA-Z]+", number.upper())
    # Объединяем все буквы в одну строку
    letters_only = "".join(letters)
    # Преобразуем русские буквы в латинские для унификации поиска
    return ru_to_lat(letters_only)

def search_numbers_by_letters(search_query, max_results=50):
    """
    Улучшенный поиск номеров по буквам в Excel файле
    """
    try:
        # Проверяем существование файла
        if not os.path.exists(EXCEL_FILE):
            logger.error(f"Excel файл не найден: {EXCEL_FILE}")
            return []
        
        # Открываем Excel файл
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Преобразуем поисковый запрос
        query = ru_to_lat(search_query.upper().strip())
        
        if not query:
            return []
        
        results = []
        
        # Проходим по всем строкам (начиная со 2-й, пропуская заголовки)
        for row in range(2, ws.max_row + 1):
            number = str(ws.cell(row=row, column=1).value or "")
            region = str(ws.cell(row=row, column=2).value or "")
            price = str(ws.cell(row=row, column=3).value or "")
            comment = str(ws.cell(row=row, column=4).value or "")
            
            # Извлекаем буквы из номера
            number_letters = extract_letters_from_number(number)
            
            # Проверяем, содержит ли номер искомые буквы
            if query in number_letters:
                # Формируем строку результата
                result_line = f"{number}"
                if region and region != "None":
                    result_line += f" (регион {region})"
                if price and price != "None":
                    result_line += f" - {price}₽"
                if comment and comment != "None":
                    result_line += f" {comment}"
                
                results.append(result_line)
                
                # Ограничиваем количество результатов
                if len(results) >= max_results:
                    break
        
        wb.close()
        return results
        
    except Exception as e:
        logger.error(f"Ошибка при поиске по буквам: {e}")
        return []

def search_numbers_by_digits(search_query, max_results=50):
    """
    Поиск номеров по цифрам в Excel файле
    """
    try:
        # Проверяем существование файла
        if not os.path.exists(EXCEL_FILE):
            logger.error(f"Excel файл не найден: {EXCEL_FILE}")
            return []
        
        # Открываем Excel файл
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        query = search_query.strip()
        
        if not query:
            return []
        
        results = []
        
        # Проходим по всем строкам (начиная со 2-й, пропуская заголовки)
        for row in range(2, ws.max_row + 1):
            number = str(ws.cell(row=row, column=1).value or "")
            region = str(ws.cell(row=row, column=2).value or "")
            price = str(ws.cell(row=row, column=3).value or "")
            comment = str(ws.cell(row=row, column=4).value or "")
            
            # Проверяем, содержит ли номер или регион искомые цифры
            full_number = f"{number}{region}"
            if query in full_number:
                # Формируем строку результата
                result_line = f"{number}"
                if region and region != "None":
                    result_line += f" (регион {region})"
                if price and price != "None":
                    result_line += f" - {price}₽"
                if comment and comment != "None":
                    result_line += f" {comment}"
                
                results.append(result_line)
                
                # Ограничиваем количество результатов
                if len(results) >= max_results:
                    break
        
        wb.close()
        return results
        
    except Exception as e:
        logger.error(f"Ошибка при поиске по цифрам: {e}")
        return []

# === Логирование ===
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# === /start ===
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = ReplyKeyboardMarkup([
        ["🔁 Старт"],
        ["\U0001F50D Поиск номера по цифрам (авто)", "\U0001F520 Поиск номера по буквам"],
        ["\U0001F6CD Мото номера", "\U0001F69B Прицеп номера"],
        ["\U0001F4CD Москва все номера", "\U0001F4CD Московская обл. все номера"],
        ["\U0001F6E0 Наши услуги", "\U0001F4DE Наш адрес и контакты"]
    ], resize_keyboard=True)

    await update.message.reply_text(
        "Добро пожаловать в компанию BlatZnak!\n"
        "Мы занимаемся продажей гос номеров и постановкой на учет.\n"
        "Выберите действие:",
        reply_markup=keyboard
    )

# === Показать файл целиком ===
async def send_full_file(update: Update, context: ContextTypes.DEFAULT_TYPE, filename: str):
    if not os.path.exists(filename):
        await update.message.reply_text("Файл с номерами не найден.")
        return
    try:
        with open(filename, "r", encoding="utf-8") as f:
            content = f.read()
            for i in range(0, len(content), 4000):
                await update.message.reply_text(content[i:i+4000])
    except Exception as e:
        logger.error(f"Ошибка при чтении файла {filename}: {e}")
        await update.message.reply_text("Ошибка при чтении файла.")

# === Универсальный обработчик ===
async def unified_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    user_data = context.user_data

    if text == "🔁 Старт":
        await start(update, context)
        return

    if user_data.get("expecting_page_size"):
        try:
            page_size = int(text)
            if page_size < 1 or page_size > 100:
                raise ValueError
            user_data["page_size"] = page_size
            user_data["expecting_page_size"] = False
            category = user_data["selected_category"]
            file_map = {
                "moscow": MOSCOW_FILE,
                "mosreg": MOSREG_FILE,
                "trailer": TRAILER_FILE
            }
            filename = file_map[category]
            if os.path.exists(filename):
                with open(filename, "r", encoding="utf-8") as f:
                    lines = f.readlines()
                for i in range(0, len(lines), page_size):
                    chunk = "".join(lines[i:i + page_size])
                    await update.message.reply_text(chunk)
            else:
                await update.message.reply_text("Файл не найден.")
        except ValueError:
            user_data["expecting_page_size"] = False
            await update.message.reply_text(
                "\u2757 Сейчас ожидалось число от 1 до 100 для показа номеров. "
                "Попробуйте ещё раз или нажмите 🔁 Старт для возврата в меню."
            )
        return

    elif text == "\U0001F520 Поиск номера по буквам":
        user_data["expecting_letter_search"] = True
        await update.message.reply_text("Введите буквы для поиска (например, СС, АА, МК):")
        return

    elif user_data.get("expecting_letter_search"):
        user_data["expecting_letter_search"] = False
        
        # Используем улучшенный алгоритм поиска по буквам
        results = search_numbers_by_letters(text, max_results=50)
        
        if results:
            reply = f"🔍 Найдено номеров с буквами '{text.upper()}': {len(results)}\n\n"
            reply += "\n".join(results)
        else:
            reply = f"❌ Номеров с буквами '{text.upper()}' не найдено."
        
        # Разбиваем длинные сообщения
        for i in range(0, len(reply), 4000):
            await update.message.reply_text(reply[i:i+4000])
        return

    elif text == "\U0001F50D Поиск номера по цифрам (авто)":
        user_data["expecting_digit_search"] = True
        await update.message.reply_text("Отправьте цифры номера для поиска (например, 777, 123):")
        return

    elif user_data.get("expecting_digit_search"):
        user_data["expecting_digit_search"] = False
        
        # Используем поиск по цифрам
        results = search_numbers_by_digits(text, max_results=50)
        
        if results:
            reply = f"🔍 Найдено номеров с цифрами '{text}': {len(results)}\n\n"
            reply += "\n".join(results)
        else:
            reply = f"❌ Номеров с цифрами '{text}' не найдено."
        
        # Разбиваем длинные сообщения
        for i in range(0, len(reply), 4000):
            await update.message.reply_text(reply[i:i+4000])
        return

    elif text == "\U0001F6CD Мото номера":
        await send_full_file(update, context, MOTO_FILE)

    elif text in {
        "\U0001F69B Прицеп номера",
        "\U0001F4CD Москва все номера",
        "\U0001F4CD Московская обл. все номера"
    }:
        category_map = {
            "\U0001F69B Прицеп номера": "trailer",
            "\U0001F4CD Москва все номера": "moscow",
            "\U0001F4CD Московская обл. все номера": "mosreg"
        }
        category = category_map[text]
        user_data["expecting_page_size"] = True
        user_data["selected_category"] = category
        await update.message.reply_text("Сколько номеров показать на странице? (например, 30)")

    elif text == "\U0001F6E0 Наши услуги":
        await update.message.reply_text(
            "\U0001F4CC Наши услуги:\n"
            "- Дубликат номеров\n"
            "- Постановка автомобиля на учет\n"
            "- Продажа красивых номеров\n"
            "- Страхование"
        )

    elif text == "\U0001F4DE Наш адрес и контакты":
        await update.message.reply_text(
            "\U0001F3E2 Адрес: улица Твардовского, 8к5с1, Москва\n"
            "\U0001F4CD [Открыть в Яндекс.Навигаторе](https://yandex.ru/navi/?ol=geo&text=%D1%83%D0%BB%D0%B8%D1%86%D0%B0%20%D0%A2%D0%B2%D0%B0%D1%80%D0%B4%D0%BE%D0%B2%D1%81%D0%BA%D0%BE%D0%B3%D0%BE,%208%D0%BA5%D1%811&sll=37.388268,55.792574)\n"
            "\u260E [Позвонить: +7 (966) 000-26-26](tel:+79660002626)\n"
            "\U0001F4AC Telegram: @blatznak77\n"
            "\U0001F4F1 [Написать в WhatsApp](https://wa.me/79660002626)",
            parse_mode="Markdown"
        )
    else:
        # Если пользователь ввел что-то другое, пытаемся найти по цифрам
        results = search_numbers_by_digits(text, max_results=50)
        
        if results:
            reply = f"🔍 Найдено номеров с '{text}': {len(results)}\n\n"
            reply += "\n".join(results)
        else:
            reply = f"❌ Номеров с '{text}' не найдено."
        
        # Разбиваем длинные сообщения
        for i in range(0, len(reply), 4000):
            await update.message.reply_text(reply[i:i+4000])

# === Main ===
def main():
    if not BOT_TOKEN:
        logger.error("BOT_TOKEN не установлен в переменных окружения!")
        return
    
    logger.info("Запуск бота...")
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, unified_handler))
    
    logger.info("Бот запущен и готов к работе!")
    app.run_polling()

if __name__ == "__main__":
    main()

